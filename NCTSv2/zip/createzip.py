from datetime import datetime
from io import BytesIO
import zipfile
import logging
import openpyxl

def write_to_excel(data):
    # Create a BytesIO buffer to hold the ZIP file data
    zip_buffer = BytesIO()

    # Create a ZIP archive in the buffer
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for entry in data:
            # Create a new workbook
            wb = openpyxl.Workbook()
            
            # Remove the default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Create first sheet for header1 and values1
            ws1 = wb.create_sheet(title="Main Data")
            
            header1 = [
                "Lloyds",
                "Verblijfs",
                "Agent",
                "Vessel/Ship",
                "POL",
                "Quay"
            ] 
            
            values1 = [
                entry.get('Loyds', ''),
                entry.get('Stay', ''),
                entry.get('Agent code', ''),
                entry.get('Vissel', ''),
                entry.get('POL', ''),
                entry.get('Quay', '')
            ]
            
            # Add header1 to the first sheet
            ws1.append(header1)
            
            # Add values1 to the first sheet
            ws1.append(values1)
            
            # Create second sheet for header2 and items
            ws2 = wb.create_sheet(title="Items Data")
            
            header2 = [
                "Container",
                "Collis",
                "Gross weight",
                "package type",
                "HS code",
                "Goods Description",
                "Article number",
                "Item",
                "BL",
            ]
            
            # Add header2 to the second sheet
            ws2.append(header2)
            
            # Add items to the second sheet
            if 'Items' in entry:
                for item in entry['Items']:
                    item_values = [
                        item.get('Container', ''),
                        item.get('Package', ''),
                        item.get('Gross', ''),
                        item.get('Package', ''),  # Assuming this field exists in your data
                        item.get('HS code', ''),
                        item.get('Description', ''),
                        item.get('Article', ''),
                        item.get('Item', ''),
                        entry.get('BLnumber', '')  # Assuming BLnumber is the BL field
                    ]
                    ws2.append(item_values)
            
            # Optionally, adjust column widths for better formatting
            for ws in [ws1, ws2]:
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
            
            # Save the file with the container name
            excel_filename = f"{entry['Container']}.xlsx"
            with BytesIO() as excel_buffer:
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                zip_file.writestr(excel_filename, excel_buffer.read())
    
    # Finalize the ZIP file and prepare it for the response
    zip_buffer.seek(0)  # Go to the beginning of the ZIP buffer
    return zip_buffer.getvalue()