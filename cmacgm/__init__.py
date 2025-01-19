import azure.functions as func
import logging
import json
import base64
import io  # For in-memory file handling
import pandas as pd
import openpyxl

from cmacgm.helpers.functions import process_container_data
from cmacgm.helpers.sentEmail import json_to_xml

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Processing file upload request.')

    # Attempt to get the JSON body from the request
    try:
        body = req.get_json()
        base64_files = body.get('files', [])
    except Exception as e:
        return func.HttpResponse(
            body=json.dumps({"error": "Invalid request format"}),
            status_code=400,
            mimetype="application/json"
        )

    if not base64_files:
        return func.HttpResponse(
            body=json.dumps({"error": "No files provided"}),
            status_code=400,
            mimetype="application/json"
        )

    extracted_data = []

    for base64_file in base64_files:
        filename = base64_file.get('filename')
        file_data = base64_file.get('file')

        if not filename or not file_data:
            continue

        # Decode the base64-encoded file
        try:
            decoded_data = base64.b64decode(file_data)
            decoded_data = decoded_data.decode('utf-8')  # Ensure it's in text format

            # Load the CSV data into a pandas DataFrame using in-memory StringIO
            try:
                csv_data = io.StringIO(decoded_data)  # In-memory file-like object for CSV
                # Specify the delimiter based on the CSV structure (likely semicolon `;`)
                df = pd.read_csv(csv_data, delimiter=';')  # Set delimiter to semicolon

                # Convert the DataFrame to an Excel file in-memory using BytesIO
                excel_data = io.BytesIO()  # In-memory binary stream for Excel
                with pd.ExcelWriter(excel_data, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)

                # Seek to the beginning of the Excel data stream
                excel_data.seek(0)

                # Load the Excel file from the in-memory stream using openpyxl
                wb = openpyxl.load_workbook(excel_data)
                sheet = wb.active  # Get the active sheet

                # Extract static values (those not in the items)
                data = {
                    "container": sheet.cell(row=2, column=8).value,
                    "Vissel": sheet.cell(row=2, column=4).value,
                    "Port Of Loading": sheet.cell(row=2, column=12).value,
                    "LoydsNumber": sheet.cell(row=2, column=2).value,
                    "BL number": sheet.cell(row=2, column=10).value,
                    "Article": sheet.cell(row=2, column=9).value,
                    "Agent Code": sheet.cell(row=2, column=1).value,
                    "Stay": sheet.cell(row=2, column=3).value,
                    "Quay": sheet.cell(row=2, column=13).value,
                }

                # Extract dynamic "items" that may exist in columns B, C, D, etc.
                items = []
                row = 2  # Start from row 2, as in your original logic

                while True:
                    # Extract dynamic values from columns (starting from column 14)
                    itemNbr = sheet.cell(row=row, column=14).value
                    description = sheet.cell(row=row, column=15).value
                    gross_weight = sheet.cell(row=row, column=16).value
                    packages = sheet.cell(row=row, column=17).value

                    # If any of the required cells are empty, break the loop
                    if not (itemNbr and description and gross_weight and packages):
                        break

                    # Create the item dictionary
                    item = {
                        "item": itemNbr,
                        "description": description,
                        "Gross Weight": gross_weight,
                        "Net Weight": gross_weight,  # Assuming net weight is the same as gross weight
                        "Packages": packages
                    }

                    items.append(item)
                    row += 1  # Move to the next row

                # If items exist, add them to the data
                if items:
                    data["items"] = items

                extracted_data.append(data)

            except Exception as e:
                logging.error(f"Error converting CSV to Excel: {e}")
                return func.HttpResponse(
                    body=json.dumps({"error": f"Failed to convert CSV to Excel: {str(e)}"}),
                    status_code=500,
                    mimetype="application/json"
                )

        except Exception as e:
            logging.error(f"Error decoding or processing file: {e}")
            return func.HttpResponse(
                body=json.dumps({"error": f"Failed to decode base64 file: {str(e)}"}),
                status_code=500,
                mimetype="application/json"
            )

    # Process the extracted data if necessary (e.g., further processing or conversion)
    processed_output = process_container_data(extracted_data)

    # Convert the processed output to XML if needed
    xml_data = json_to_xml(processed_output)

    # For now, returning the JSON directly
    return func.HttpResponse(
        body=xml_data,  # Or xml_data if XML conversion is needed
        status_code=200,
        mimetype="application/xml"
    )