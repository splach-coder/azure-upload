import azure.functions as func
import logging
import json
import os
import openpyxl
import base64

from bbl.helpers.functions import process_container_data
from bbl.helpers.sentEmail import json_to_xml

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Processing file upload request.')

    # Attempt to get the JSON body from the request
    try:
        body = req.get_json()
        base64_files = body.get('files', [])
    except Exception as e:
        return func.HttpResponse(
            body=json.dumps({"error": "Invalid request format"}),
            status_code=400,
            mimetype="application/json"
        )

    if not base64_files:
        return func.HttpResponse(
            body=json.dumps({"error": "No files provided"}),
            status_code=400,
            mimetype="application/json"
        )

    extracted_data = []

    for base64_file in base64_files:
        filename = base64_file.get('filename')
        file_data = base64_file.get('file')

        if not filename or not file_data:
            continue

        # Decode the base64-encoded file
        try:
            decoded_data = base64.b64decode(file_data)

            # Save the uploaded file temporarily
            uploaded_file_path = f'./temp/{filename}'
            with open(uploaded_file_path, 'wb') as temp_file:
                temp_file.write(decoded_data)

            # Process the Excel file using openpyxl
            try:
                wb = openpyxl.load_workbook(uploaded_file_path)
                sheet = wb.active  # Get the active sheet

                # Extract static values (those not in the items)
                data = {
                    "container": sheet.cell(row=1, column=1).value,
                    "Incoterm": sheet.cell(row=6, column=2).value,
                    "Freight": sheet.cell(row=13, column=2).value,
                    "Vat 1": sheet.cell(row=14, column=2).value,
                    "Vat 2": sheet.cell(row=15, column=2).value,
                }

                # Extract dynamic "items" that may exist in columns B, C, D, etc.
                items = []
                column = 2  # Start from column B (2), and check up to column D, E, etc.

                while True:
                    # Check the required cell values for emptiness
                    hscode = sheet.cell(row=7, column=column).value
                    valeur = sheet.cell(row=8, column=column).value
                    devises = sheet.cell(row=9, column=column).value
                    gross_weight = sheet.cell(row=10, column=column).value
                    net_weight = sheet.cell(row=11, column=column).value
                    packages = sheet.cell(row=12, column=column).value

                    # If any of the required cells are empty, break the loop
                    if not (hscode and valeur and devises and gross_weight and net_weight and packages):
                        break
                    
                    # Create the item dictionary
                    item = {
                        "HSCODE": hscode,
                        "VALEUR": valeur,
                        "DEVISES": devises,
                        "Gross Weight": gross_weight,
                        "Net Weight": net_weight,
                        "Packages": packages
                    }

                    items.append(item)
                    column += 1  # Move to the next column

                # If items exist, add them to the data
                if items:
                    data["items"] = items

                extracted_data.append(data)

            except Exception as e:
                logging.error(f"Error processing Excel file: {e}")
                return func.HttpResponse(
                    body=json.dumps({"error": f"Failed to process Excel file: {str(e)}"}),
                    status_code=500,
                    mimetype="application/json"
                )

            processed_output = process_container_data(extracted_data)

            xml_data = json_to_xml(processed_output)

            # Delete the temporary uploaded file
            os.remove(uploaded_file_path)

        except Exception as e:
            logging.error(f"Error decoding or saving file: {e}")
            return func.HttpResponse(
                body=json.dumps({"error": f"Failed to decode base64 file: {str(e)}"}),
                status_code=500,
                mimetype="application/json"
            )

    # Construct the JSON response with the extracted data
    response_body = xml_data

    return func.HttpResponse(
        body=response_body,
        status_code=200,
        mimetype="application/json"
    )
