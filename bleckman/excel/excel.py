import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

def create_excel(data):
    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"
    
    incoterm = data.get("Incoterm", "")
    inco, term = ['', '']
    if incoterm and len(incoterm) > 1:
        inco, term = incoterm
        
    address = data.get("Address", "")
    name, street, city, postal, country = ['', '', '', '', '']
    if address and len(address) == 1:
        address = address[0]
        if len(address) == 5:
            name, street, city, postal, country = address#.get("Company name", ""), address.get("Street", ""), address.get("City", ""), address.get("Postal Code", ""), address.get("Country", ""),

    # Data for the labels
    datat = [
        ["Instruction 1", data.get("Instruction 1", "")],
        ["ILS Number", data.get("ILS_NUMBER", "")],
        ["Exit Office", data.get("Exit Office", "")],
        ["Site", "KR"],
        ["EUR1", "XXXXX"],
        ["Inco", inco, term],
        ["id Bleckmann", data.get("Reference", "")],
        ["Ordernr", "XXXXXXX"],
        ["Date", data.get("Inv Date", "")],
        ["Colli", "", "", "", "", "", "", ""],
        ["Weight", "", "", "", "", "", "", ""],
        ["Currency", data.get("Currency", "")],
        ["", "", "", "", "", "", "", ""],
        ["Colli - weight", "", "", "", "", "", "", ""],
        ["Col #1", "Col #2", "Col #3", "Col #4", "Col #5", "", "", ""],
        [data.get("collis", ""), data.get("grosses", ""), "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", ""],
        ["Consignee", "", "", "", "", "", "", ""],
        ["Col #1", "Col #2", "Col #3", "Col #4", "Col #5"],
        [name, street, city, postal, country],
        ["", "", "", "", "", "", "", ""],
        ["Invoicenumber", "", "", "", "", "", "", ""],
        ["Col #1", "Col #2", "", "", "", "", "", ""],
        data.get("Invoices", ""),
        ["", "", "", "", "", "", "", ""],
        ["InvoiceValue", "", "", "", "", "", "", ""],
        ["Col #1", "Col #2", "", "", "", "", "", ""],
        data.get("Totals", ""),
        ["",  "", "", "", "", "", "", ""],
        ["Items", "", "", "", "", "", "", ""],
        ["Col #1", "Col #3", "Col #3", "Col #4", "Col #5"]
    ]
    
    Items = data.get("Items", [])
    
    # Define the desired key order
    desired_order = ['Pieces', 'HS code', 'Origin', 'Price']
    
    if Items: 
        for item in Items:
            res = []
            for key in desired_order:
                res.append(item.get(key))  # Use .get() to avoid KeyError if a key is missing
            datat.append(res)         

    # Apply data to the worksheet
    for row_idx, row_data in enumerate(datat, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Apply font to all cells
            cell.font = Font(bold=False, size=11)
      
    ws["A1"].font = Font(bold=True)  
    ws["A2"].font = Font(bold=True) 
    ws["A3"].font = Font(bold=True)        
    ws["A4"].font = Font(bold=True)        
    ws["A5"].font = Font(bold=True)        
    ws["A6"].font = Font(bold=True)        
    ws["A7"].font = Font(bold=True)        
    ws["A8"].font = Font(bold=True)        
    ws["A9"].font = Font(bold=True)        
    ws["A10"].font = Font(bold=True)        
    ws["A11"].font = Font(bold=True)        
    ws["A12"].font = Font(bold=True)        
    ws["A14"].font = Font(bold=True)        
    ws["A15"].font = Font(bold=True)        
    ws["B15"].font = Font(bold=True)        
    ws["C15"].font = Font(bold=True)        
    ws["D15"].font = Font(bold=True)        
    ws["E15"].font = Font(bold=True)        
    ws["A18"].font = Font(bold=True)        
    ws["A19"].font = Font(bold=True)        
    ws["B19"].font = Font(bold=True)        
    ws["C19"].font = Font(bold=True)        
    ws["D19"].font = Font(bold=True)        
    ws["E19"].font = Font(bold=True)        
    ws["A22"].font = Font(bold=True)        
    ws["A23"].font = Font(bold=True)        
    ws["B23"].font = Font(bold=True) 
           
    ws["A26"].font = Font(bold=True)        
    ws["A27"].font = Font(bold=True)        
    ws["B27"].font = Font(bold=True)
    ws["A29"].font = Font(bold=True)
            
    ws["A30"].font = Font(bold=True)        
    ws["A31"].font = Font(bold=True)        
    ws["B31"].font = Font(bold=True)        
    ws["C31"].font = Font(bold=True)        
    ws["D31"].font = Font(bold=True)        
    ws["E31"].font = Font(bold=True)       

    # Save the workbook to a BytesIO object (in memory)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream
