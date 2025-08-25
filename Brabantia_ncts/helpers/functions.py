import re
import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from io import BytesIO


# Helper function to safely convert values to float
def safe_float_conversion(value):
    if value is None:
        return 0.0  # Default to 0 if the value is None
    if isinstance(value, (int, float)):  # If it's already a number, return as is
        return float(value)
    try:
        return float(value.replace(",", "."))  # Try to replace and convert
    except (ValueError, AttributeError):
        print(f"Error converting value: {value}")
        return 0.0  # Handle conversion error, default to 0 or other logic
    
    
def merge_items_with_mrn(data):
    header_lookup = {entry["Code"]: entry["Number"] for entry in data.get("header", [])}
   
    for item in data.get("items", []):
        key = item.get("merged_EX_A_D")
        if key and key in header_lookup:
            item["MRN_number"] = header_lookup[key]
    
    # Sort items by merged_EX_A_D in alphabetical order (A-Z)
    # Handle None values by placing them at the end
    data["items"] = sorted(
        data.get("items", []), 
        key=lambda x: x.get("merged_EX_A_D") or ""
    )
   
    return data

import openpyxl
from openpyxl.styles import Font
from io import BytesIO
import uuid
from datetime import datetime

def create_excel_from_merged_data(merged_data, filename_prefix="export"):
    """
    Creates an Excel file with one sheet matching the structure of the "cls_riccverf" sheet
    from the provided merged data and returns it as a BytesIO object with a unique filename.
    
    Args:
        merged_data (dict): The merged data containing items and header information
        filename_prefix (str): Optional prefix for the filename
        
    Returns:
        tuple: (BytesIO file stream, unique filename)
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1" 
    
    # Define the headers based on the original structure
    headers = [
        "Article", "Article description", "Stat.no.", "Preferential EU origin",
        "Production country", "Pieces", "Packs", "Currency", "Price", "Amount",
        "Net weight [kg]", "Gross weight [kg]", "Prod. unit", "EX A / EX D",
        "Factuur", "Merged Factuur Number", "MRN_number"
    ]
    
    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Write data rows
    for row_num, item in enumerate(merged_data.get("items", []), 2):
        ws.cell(row=row_num, column=1, value=item.get("Article"))
        ws.cell(row=row_num, column=2, value=item.get("Description"))
        ws.cell(row=row_num, column=3, value=item.get("Stat No"))
        ws.cell(row=row_num, column=4, value=item.get("Preferential EU Origin"))
        ws.cell(row=row_num, column=5, value=item.get("Production Country"))
        ws.cell(row=row_num, column=6, value=item.get("Pieces"))
        ws.cell(row=row_num, column=7, value=item.get("Packs"))
        ws.cell(row=row_num, column=8, value=item.get("Currency"))
        ws.cell(row=row_num, column=9, value=item.get("Price"))
        ws.cell(row=row_num, column=10, value=item.get("Amount"))
        ws.cell(row=row_num, column=11, value=item.get("Net Weight"))
        ws.cell(row=row_num, column=12, value=item.get("Gross Weight"))
        ws.cell(row=row_num, column=13, value=item.get("Production Unit"))
        ws.cell(row=row_num, column=14, value=item.get("EX A / EX D"))
        ws.cell(row=row_num, column=15, value=item.get("Invoice"))
        ws.cell(row=row_num, column=16, value=item.get("merged_EX_A_D"))
        ws.cell(row=row_num, column=17, value=item.get("MRN_number"))
        

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Generate unique filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = str(uuid.uuid4())[:8]
    filename = f"{filename_prefix}_{timestamp}_{unique_id}.xlsx"

    # Save to BytesIO
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream, filename
