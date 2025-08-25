import azure.functions as func
import logging
import json
import os
import openpyxl
import base64

from Brabantia_ncts.helpers.functions import create_excel_from_merged_data, merge_items_with_mrn, safe_float_conversion

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Processing file upload request.')

    try:
        body = req.get_json()
        base64_files = body.get('files', [])

    except Exception:
        return func.HttpResponse(
            json.dumps({"error": "Invalid request format"}),
            status_code=400,
            mimetype="application/json"
        )

    if not base64_files:
        return func.HttpResponse(
            json.dumps({"error": "No files provided"}),
            status_code=400,
            mimetype="application/json"
        )

    extracted_data = {
        "items": [],
        "header": []
    }

    for base64_file in base64_files:
        filename = base64_file.get('filename')
        file_data = base64_file.get('file')

        if not filename or not file_data:
            continue

        try:
            decoded_data = base64.b64decode(file_data)
            temp_dir = os.getenv('TEMP', '/tmp')
            file_path = os.path.join(temp_dir, filename)

            with open(file_path, 'wb') as temp_file:
                temp_file.write(decoded_data)

            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_cls = None
            sheet_totalen = None

            for sheet_name in wb.sheetnames:
                if sheet_name.lower().startswith("cls"):
                    sheet_cls = wb[sheet_name]
                elif sheet_name.lower().startswith("totalen"):
                    sheet_totalen = wb[sheet_name]

            # Extract items from "cls" sheet
            if sheet_cls:
                row = 2
                while True:
                    article = sheet_cls.cell(row=row, column=1).value
                    description = sheet_cls.cell(row=row, column=2).value
                    stat_no = sheet_cls.cell(row=row, column=3).value
                    eu_origin = sheet_cls.cell(row=row, column=4).value
                    prod_country = sheet_cls.cell(row=row, column=5).value
                    pieces = sheet_cls.cell(row=row, column=6).value
                    packs = sheet_cls.cell(row=row, column=7).value
                    currency = sheet_cls.cell(row=row, column=8).value
                    price = sheet_cls.cell(row=row, column=9).value
                    amount = sheet_cls.cell(row=row, column=10).value
                    net_weight = sheet_cls.cell(row=row, column=11).value
                    gross_weight = sheet_cls.cell(row=row, column=12).value
                    prod_unit = sheet_cls.cell(row=row, column=13).value
                    ex_a_d = sheet_cls.cell(row=row, column=14).value
                    invoice = sheet_cls.cell(row=row, column=15).value

                    if not all([article, description, stat_no, eu_origin, prod_country, pieces, packs, currency, price, amount, net_weight, gross_weight, prod_unit, ex_a_d, invoice]):
                        break

                    extracted_data["items"].append({
                        "Article": article,
                        "Description": description,
                        "Stat No": stat_no,
                        "Preferential EU Origin": eu_origin,
                        "Production Country": prod_country,
                        "Pieces": pieces,
                        "Packs": packs,
                        "Currency": currency,
                        "Price": safe_float_conversion(price),
                        "Amount": safe_float_conversion(amount),
                        "Net Weight": safe_float_conversion(net_weight),
                        "Gross Weight": safe_float_conversion(gross_weight),
                        "Production Unit": prod_unit,
                        "EX A / EX D": ex_a_d,
                        "Invoice": invoice,
                        "merged_EX_A_D": f"{invoice}{ex_a_d}" if ex_a_d and invoice else None
                    })
                    row += 1

            # Extract header from "Totalen" sheet
            if sheet_totalen:
                row = 4
                mergedLetter = ['A', 'D']
                mergedLetterIndex = 0
                while True:
                    col_g = sheet_totalen.cell(row=row, column=7).value  # G
                    col_h = sheet_totalen.cell(row=row, column=8).value  # H

                    if not col_g and not col_h:
                        break

                    extracted_data["header"].append({
                        "Code": f"{col_g}{mergedLetter[mergedLetterIndex]}",
                        "Number": col_h.replace('\n', ' ').strip() if col_h else None,
                    })
                    
                    if mergedLetterIndex == 0:
                        mergedLetterIndex = 1
                    else:
                        mergedLetterIndex = 0
                             
                    row += 1

            os.remove(file_path)

        except Exception as e:
            logging.error(f"Processing error: {e}")
            return func.HttpResponse(
                json.dumps({"error": str(e)}),
                status_code=500,
                mimetype="application/json"
            )

    # Merge items with MRN numbers
    merged_data = merge_items_with_mrn(extracted_data)
    
    logging.error(json.dumps(merged_data, indent=4))

    # Create the Excel file with unique filename
    excel_stream, filename = create_excel_from_merged_data(merged_data, "brabantia_export")
    
    # Return the Excel file as a response
    return func.HttpResponse(
        excel_stream.getvalue(),
        status_code=200,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Allow-Origin": "*"
        }
    )
