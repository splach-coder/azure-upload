import base64
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from AI_agents.Mistral.MistralDocumentQA import MistralDocumentQA

def extract_clean_excel_from_pdf(base64_pdf: str, filename) -> BytesIO:
    prompt = (
        "Extract all table data from the PDF text into a clean JSON object.\n\n"
        "- Rows with **two stars \"**\" in a specific column mark a Subtotal row.\n"
        "- Rows with **three stars \"***\" and no other text in that column mark the Grand Total row.\n"
        "- Subtotal rows summarize the section above and include an extra number on the right side called \"Collis\".\n"
        "- Grand Total summarizes the entire table.\n"
        "- Remove repeated headers like \"Customs code\".\n"
        "- Preserve \"Bill. Doc.\" and \"Comm. Code\" columns as strings.\n"
        "- Return only a JSON object with a \"rows\" array.\n"
        "- Mark subtotals with \"SubTotal\": true and add \"Collis\".\n"
        "- Keep row order.\n"
        "- Remove empty or irrelevant rows.\n"
        "- Include references like \"Reference\" if found.\n\n"
        "**IMPORTANT:** Return ONLY the JSON object with an array named \"rows\". No explanations, no text, no markdown, no formatting."
    )

    # Mistral call
    qa = MistralDocumentQA()
    response = qa.ask_document(base64_pdf, prompt, filename=filename)

    # Clean response
    raw = response.replace("```", "").replace("json", "").strip()
    parsed = json.loads(raw)

    # Clean & filter
    cleaned = [
        {k: v for k, v in row.items() if k != "WeightUnit"}
        for row in parsed["rows"]
        if "Comm. Code" in row
    ]

    headers = list({k for row in cleaned for k in row.keys()})

    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(cleaned, start=2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(header), 15)

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    return output_stream