import uuid
import azure.functions as func
import logging
import json
import os
import openpyxl
import base64
import re

from AI_agents.OpenAI.custom_call import CustomCall
from VanPoppel_BlackEnDecker.excel.create_excel import write_to_excel
from VanPoppel_BlackEnDecker.functions.functions import extract_email_body


def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Processing file upload request.')

    # Attempt to get the JSON body from the request
    try:
        body = req.get_json()
        excels = body.get('files', [])
        email = body.get('email', [])
        subject = body.get('subject', [])

    except Exception as e:
        return func.HttpResponse(
            body=json.dumps({"error": "Invalid request format"}),
            status_code=400,
            mimetype="application/json"
        )

    if not excels:
        return func.HttpResponse(
            body=json.dumps({"error": "No files provided"}),
            status_code=400,
            mimetype="application/json"
        )

    extracted_data_from_excels = []
    TotalPrice = 0

    # Helper functions
    def _normalize_header(h):
        if h is None:
            return ""
        s = str(h).strip().lower()
        s = re.sub(r'[^a-z0-9]', ' ', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    def _safe_get(row, idx):
        try:
            if idx is None:
                return None
            return row[idx] if idx < len(row) else None
        except Exception:
            return None

    def _parse_number(val):
        try:
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).strip()
            if s == '':
                return 0.0
            s = s.replace('\u00A0', '').replace('\xa0', '')
            s_clean = re.sub(r'[^\d\.,\-]', '', s)
            if s_clean == '' or s_clean == '-' or s_clean == '--':
                return 0.0
            if '.' in s_clean and ',' in s_clean:
                if s_clean.find(',') < s_clean.find('.'):
                    s_clean = s_clean.replace(',', '')
                else:
                    s_clean = s_clean.replace('.', '').replace(',', '.')
            elif ',' in s_clean and '.' not in s_clean:
                parts = s_clean.split(',')
                if len(parts[-1]) == 3:
                    s_clean = s_clean.replace(',', '')
                else:
                    s_clean = s_clean.replace(',', '.')
            if s_clean in ['', '-', '--']:
                return 0.0
            return float(s_clean)
        except Exception:
            try:
                return float(str(val).replace(',', '').strip())
            except Exception:
                return 0.0

    # Process each uploaded Excel
    for excel in excels:
        filename = excel.get('filename')
        file_data = excel.get('file')

        if not filename or not file_data:
            continue

        try:
            decoded_data = base64.b64decode(file_data)
            temp_dir = os.getenv('TEMP', '/tmp')
            uploaded_file_path = os.path.join(temp_dir, filename)

            with open(uploaded_file_path, 'wb') as temp_file:
                temp_file.write(decoded_data)

            workbook = openpyxl.load_workbook(uploaded_file_path)
            sheet = workbook.active

            # --- Dynamic header detection ---
            header_tuple = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            header_norm = [_normalize_header(h) for h in header_tuple] if header_tuple else []

            net_idx = None
            gross_idx = None
            for idx, hn in enumerate(header_norm):
                if hn and 'net' in hn and ('weight' in hn or 'wt' in hn or 'kg' in hn or 'kgs' in hn):
                    net_idx = idx
                if hn and 'gross' in hn and ('weight' in hn or 'wt' in hn or 'kg' in hn or 'kgs' in hn):
                    gross_idx = idx

            # Looser fallback
            if net_idx is None:
                for idx, hn in enumerate(header_norm):
                    if hn and (hn == 'net' or hn.startswith('net ') or ' net ' in f' {hn} '):
                        net_idx = idx
                        break
            if gross_idx is None:
                for idx, hn in enumerate(header_norm):
                    if hn and (hn == 'gross' or hn.startswith('gross ') or ' gross ' in f' {hn} '):
                        gross_idx = idx
                        break

            if net_idx is None:
                net_idx = 10
            if gross_idx is None:
                gross_idx = 13

            items = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    break

                BillingDoc = row[0] if len(row) > 0 else None
                Material = row[1] if len(row) > 1 else None
                Description = row[2] if len(row) > 2 else None
                Country_of_origin = row[3] if len(row) > 3 else None
                Qty = row[4] if len(row) > 4 else 0
                TotalCost = row[6] if len(row) > 6 else 0
                Currency = row[7] if len(row) > 7 else ''
                Incoterm = row[8] if len(row) > 8 else ''
                Commodity = row[9] if len(row) > 9 else ''

                Net_wt_raw = _safe_get(row, net_idx)
                Gross_wt_raw = _safe_get(row, gross_idx)
                Net_wt = _parse_number(Net_wt_raw)
                Gross_wt = _parse_number(Gross_wt_raw)

                if not BillingDoc or not Material or not Description:
                    continue

                item = {
                    "Invoice No": str(BillingDoc),
                    'Incoterm': str(Incoterm),
                    "Material": str(Material),
                    "Commodity Code": str(Commodity),
                    "Description": str(Description),
                    "Country of Origin": str(Country_of_origin),
                    "Qty": int(Qty) if str(Qty).isdigit() else 0,
                    "Total Cost": float(TotalCost) if TotalCost else 0.00,
                    "Value": float(str(TotalCost).replace(',', '')) if TotalCost else 0.00,
                    "Currency": str(Currency),
                    "Net Wt": float(Net_wt) if Net_wt else 0.00,
                    "Gross Wt": float(Gross_wt) if Gross_wt else 0.00
                }

                TotalPrice += item["Total Cost"]
                items.append(item)

            extracted_data_from_excels.append(items)
            os.remove(uploaded_file_path)

        except Exception as e:
            return func.HttpResponse(
                body=json.dumps({"error": str(e)}, indent=4),
                status_code=500,
                mimetype="application/json"
            )

    # Process email extraction
    call = CustomCall()
    email = extract_email_body(email)
    role = "you are a data extraction agent. Your task is to extract specific fields from the email text and return them in JSON format. The fields include ShipmentReference, NetWeight, GrossWeight, Incoterm, FreightCost, Collis, OfficeOfExit, and PlaceOfDelivery. If a field is missing, return its value as null. Use number format for numerical values (no units like KG or €). Format 'Incoterm+place name' as a simple lowercase string like 'fca shanghai'."
    prompt = f"""
    Extract the following fields from the email text and return the output as pure JSON with no additional text or formatting. If a field is missing, return its value as null. Use number format for numerical values (no units like KG or €). Format "Incoterm+place name" as a simple lowercase string like "fca shanghai".

    Fields to extract:
    - ShipmentReference
    - NetWeight (as number)
    - GrossWeight (as number)
    - Incoterm (as lowercase string, include place name, e.g., "fca marseille")
    - FreightCost (as an object: {{ "value": number, "currency": string }})
    - Collis (as number)
    - OfficeOfExit (as text)
    - PlaceOfDelivery (as object: {{ "company_name": string, "street": string, "city": string, "postal_code": string, "country_code": string (2-letter code) }})

    Email text:
    \"\"\"
    {email}
    \"\"\"
    """

    email_data = call.send_request(role, prompt)
    email_data = email_data.replace("```", "").replace("json", "").strip()
    email = json.loads(email_data)

    result_data = {**email, "Total Value": TotalPrice, "Items": extracted_data_from_excels[0]}

    try:
        excel_file = write_to_excel(result_data)
        logging.info("Generated Excel file.")

        reference = result_data.get("ShipmentReference", "")
        if not reference:
            reference = f"ref-{uuid.uuid4().hex}"

        headers = {
            'Content-Disposition': 'attachment; filename="' + reference + '.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }

        return func.HttpResponse(excel_file.getvalue(), headers=headers, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        logging.error(f"Error: {e}")
        return func.HttpResponse(
            f"Error processing request: {e}", status_code=500
        )
