from io import BytesIO
import json
from openpyxl.styles import Font, PatternFill
import re
import openpyxl

def extract_currency_symbol(cell):
    """
    Extract the currency symbol from an Excel cell's number format.
    """
    # Define known currency symbols
    currency_symbols = {
        '€': 'Euro',
        '$': 'Dollar',
        '£': 'Pound',
        '¥': 'Yen',
        '₹': 'Rupee',
        '₩': 'Won',
        '₽': 'Ruble',
        '₺': 'Lira',
        '₨': 'Rupee',
        '₦': 'Naira',
        '฿': 'Baht',
        '₫': 'Dong'
    }

    # Extract the number format from the cell
    currency_format = cell.number_format if cell is not None else ""

    # Find the currency symbol
    for symbol in currency_symbols:
        if symbol in currency_format:
            return symbol

    return ''  # Default if no known currency symbol is found

def clean_string(input_string):
    # Use regex to match only alphanumeric characters
    cleaned_string = re.sub(r'[^a-zA-Z0-9]', '', input_string)
    return cleaned_string    

def extract_valid_container(container_string):
    print(container_string)
    if container_string:
        container = clean_string(container_string)

        # Check if the container matches the format 4 chars and 7 digits
        pattern = r'^[A-Z]{4}\d{7}$'
        if not re.match(pattern, container):
            return False

        return container
    else :
        return False
    
def write_to_excel(json_string):
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    data = json_string

    row_keys = []
    row_values = []
    row_empty = []

    rows_data = []

    header = [
        "Invoicenumber",
        "Client",
        "Commodity",
        "Partnumber",
        "Description",
        "Packaging",
        "Origin",
        "Invoiced",
        "Unit",
        "Pieces",
        "Gross",
        "Net",
        "EXW value",
        "Cif Value",
        "Fob Value"
    ]

    for key, value in data.items():
        # Handle array values
        if key == "items":
            #logic here
            for obj in value:
                mini_row = []
                for key2, value2 in obj.items():
                    mini_row.append(value2)
                rows_data.append(mini_row)
        else:
            row_keys.append(key)
            row_values.append(value)
            row_empty.append("")

    # Add keys (headers) to the first row
    ws.append(row_keys)

    # Add values to the second row
    ws.append(row_values)

        # Add values to the second row
    ws.append(row_empty)

    # Add values to the second row
    ws.append(header)

    for arr in rows_data:
        ws.append(arr)

    # Optionally, adjust column widths for better formatting
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Iterate through rows with enumerate to get the row index
    for i, row in enumerate(ws.iter_rows(), start=1):  # Use ws.iter_rows() and start from 1
        # Make the first row bold
        if i == 1:
            for cell in row:
                cell.font = Font(bold=True)
        
        # Fill the fourth row with a yellow background color
        if i == 4:
            for cell in row:
                if cell.value :
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save the workbook to a BytesIO object (in memory)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream

def format_float_values(data):
    for item in data:
        for key, value in item.items():
            if isinstance(value, float):
                item[key] = round(value, 2)  # Round to two decimal places
    return data

def find_value_and_get_next(sheet, search_value, start_column=2, step=1, direction='right', return_type='cell'):
    """
    Searches for a value in a specified column and retrieves cell objects a specified
    number of steps away. Returns up to two cells, prioritizing the first non-empty cell.

    Parameters:
    - sheet (Worksheet): The worksheet object.
    - search_value (str): The value to search for.
    - start_column (int): The column number to search in (default is 2).
    - step (int): The number of cells to skip from the found cell (default is 1, i.e., the next cell).
    - direction (str): 'down' to get cells below, 'right' for cells to the right.
                       Can also be 'up' or 'left' for opposite directions.
    - return_type (str): "cell" (default) to return the cell object, or "value" to return the cell's value.

    Returns:
    - The first non-empty cell if found; if the first cell is empty, returns the second cell.
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    found_cells = []

    # Search for cells containing the specified value in the start_column
    for row in range(1, max_row + 1):
        cell = sheet.cell(row=row, column=start_column)
        if cell.value == search_value:
            # Determine target cell based on direction and step
            if direction == 'down':
                target_row = row + step
                if target_row <= max_row:
                    target_cell = sheet.cell(row=target_row, column=start_column)
            elif direction == 'up':
                target_row = row - step
                if target_row >= 1:
                    target_cell = sheet.cell(row=target_row, column=start_column)
            elif direction == 'right':
                target_column = start_column + step
                if target_column <= max_col:
                    target_cell = sheet.cell(row=row, column=target_column)
            elif direction == 'left':
                target_column = start_column - step
                if target_column >= 1:
                    target_cell = sheet.cell(row=row, column=target_column)
            else:
                return ""  # Invalid direction

            # Add the target cell to found_cells if it exists
            if target_cell:
                found_cells.append(target_cell)

            # Stop after finding two cells
            if len(found_cells) == 2:
                break

    # Return the first non-empty cell; if the first is empty, return the second
    if found_cells:
        return found_cells[0] if found_cells[0].value else (found_cells[1] if len(found_cells) > 1 else "")
    else:
        return ""

def get_value_with_search(sheet, search_term, search_column=2, offset=3, direction='right', return_type='value'):
    """
    Searches for `search_term` in the specified column and returns the cell or its value
    at a specified offset. If not found or if the found cell is empty, returns an empty string.

    Parameters:
    - sheet (Worksheet): The worksheet object.
    - search_term (str): The text to search for.
    - search_column (int): The column to search in, default is 2.
    - offset (int): The number of cells away from the found cell to retrieve.
    - direction (str): Direction of offset ('right' or 'down').
    - return_type (str): "value" (default) to return the cell's value, or "cell" to return the cell object.

    Returns:
    - The value or cell object based on return_type if found, otherwise an empty string.
    """
    # Attempt to find and get the target cell based on the search term
    found_cell = find_value_and_get_next(sheet, search_term, start_column=search_column, step=offset, direction=direction, return_type='cell')

    print(found_cell)
    
    # Check if found_cell has a non-empty value and return based on return_type
    if found_cell and found_cell.value not in [None, ""]:
        return found_cell.value if return_type == 'value' else found_cell
    else:
        # Return empty string if no value found
        return "" if return_type == 'value' else None

def handle_invoice_Value(sheet):
    first_option = get_value_with_search(sheet, "22.2 Factuurwaarde")
    second_option = get_value_with_search(sheet, "22.1 Factuurwaarde")

    if (first_option) :
        return first_option
    elif (second_option) :
        return second_option
    else :
        ""

def handle_invoice_Curency(sheet):
    first_option = get_value_with_search(sheet, "22.2 Factuurwaarde",  return_type='cell')
    second_option = get_value_with_search(sheet, "22.1 Factuurwaarde",  return_type='cell')

    if first_option:
        return extract_currency_symbol(first_option)
    elif second_option:
        return extract_currency_symbol(second_option)
    else :
        "" 

def find_wy_ref(sheet):
    # Define the regex pattern for "WY" followed by 6 to 7 digits
    pattern = r"^WY\d{6,7}$"
    
    # Loop over the specified range in column E (from row 1 to row 32)
    for row in range(1, 33):
        cell_value = sheet[f"E{row}"].value  # Access the value in column E
        
        # Check if the cell value matches the pattern
        if cell_value and re.match(pattern, str(cell_value)):
            return cell_value  # Return the matched value and exit the function

    # If no match is found, return None
    return ""


