import azure.functions as func
import logging
import json
import os
import openpyxl
import base64

from Wynn.helpers.functions import find_wy_ref, extract_valid_container, format_float_values, get_value_with_search, write_to_excel, handle_invoice_Curency, handle_invoice_Value
from Wynn.helpers.adressExtractor import find_address_data

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Processing file upload request.')

    # Attempt to get the JSON body from the request
    try:
        body = req.get_json()
        base64_files = body.get('files', [])
    except Exception as e:
        return func.HttpResponse(
            body=json.dumps({"error": "Invalid request format"}),
            status_code=400,
            mimetype="application/json"
        )

    if not base64_files:
        return func.HttpResponse(
            body=json.dumps({"error": "No files provided"}),
            status_code=400,
            mimetype="application/json"
        )

    for base64_file in base64_files:
        filename = base64_file.get('filename')
        file_data = base64_file.get('file')

        if not filename or not file_data:
            continue

        # Decode the base64-encoded file
        try:
            decoded_data = base64.b64decode(file_data)

            temp_dir = os.getenv('TEMP', '/tmp')
            uploaded_file_path = os.path.join(temp_dir, filename)

            # Write the file to the temporary path
            with open(uploaded_file_path, 'wb') as temp_file:
                temp_file.write(decoded_data)

            # Process the Excel file using openpyxl
            try:
                wb = openpyxl.load_workbook(uploaded_file_path, data_only=True)
                sheet = wb.active  # Get the active sheet

                container = extract_valid_container(get_value_with_search(sheet, "31.4 Containernrs."))
                print(container)
                ref = get_value_with_search(sheet, "7 Referentienr." )
                if ref == "":
                    ref = find_wy_ref(sheet)

                # Find and parse address data
                address_data = find_address_data(sheet)

                company = address_data['company']
                street = address_data['street']
                city = address_data['city']
                country = address_data['country']

                # Define the data extraction dictionary
                data = {
                    "Goods Validation office": get_value_with_search(sheet, "54 plaats van aangifte" ),
                    "office of validation": get_value_with_search(sheet, "A Kantoor van validatie" ),
                    "Goods Location": get_value_with_search(sheet, "30 plaats goederen",),
                    "Consignee Name": company if company else "",
                    "Consignee street": street if street else "",
                    "Consignee city": city if city else "",
                    "Consignee country": country if country else "",
                    "Reference": ref,
                    "Inco Term": get_value_with_search(sheet, "20 Leveringsvoorwaarde" ),
                    "Inco Term Place": get_value_with_search(sheet, "20 Leveringsvoorwaarde", offset=4),
                    "Exit Office": get_value_with_search(sheet, "29 Kantoor van uitgang: " ),
                    "e-AD": get_value_with_search(sheet, "e-AD",  offset=2),
                    "Marks and numbers": get_value_with_search(sheet, "31.3 Merken en nummers"),
                    "transport identity departure": get_value_with_search(sheet, "18.1 Identiteit vervoermiddel vertrek" ),
                    "transport identity border": get_value_with_search(sheet, "21.1 Identiteit vervoermiddel grens" ),
                    "Invoice Value": handle_invoice_Value(sheet),
                    "Invoice value currency": handle_invoice_Curency(sheet),
                    "Transportmode border": get_value_with_search(sheet, "25 Vervoerswijze grens" ),
                    "Transportmode inland": get_value_with_search(sheet, "26 Binnenlandse vervoerswijze" ),
                    "Freight cost SEA": get_value_with_search(sheet, "25 Vervoerswijze grens", offset=5),
                    "Freight cost ROAD": get_value_with_search(sheet, "26 Binnenlandse vervoerswijze", offset=5),
                    "Containernumber": container if container else "",
                    "Seals": get_value_with_search(sheet, "D Zegels")
                }

                # Search for the specific header text in the first column to find the start of the table
                header_text = "Factuurnr"  # Replace this with your actual header text
                start_row = None

                # Iterate over the rows to find the header text in the first column
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=2).value
                    if cell_value == header_text:
                        start_row = row
                        break
                    
                if start_row is None:
                    raise ValueError("Header text not found in the first column.")

                # Extract the headers from the start_row (assuming the headers are in columns B to P)
                headers = [sheet.cell(row=start_row, column=col).value for col in range(2, 17)]  # Columns B (2) to P (16)

                # Initialize a list to hold the extracted data
                extracted_data = []

                # Counter for consecutive empty rows
                empty_row_count = 0

                # Step 2: Loop through each row after the header
                for row in range(start_row + 1, sheet.max_row + 1):
                    # Extract the data for columns B to P
                    row_data = [sheet.cell(row=row, column=col).value for col in range(2, 17)]

                    # Check if the row is considered empty (first 5 cells are None or empty)
                    if all(cell is None or str(cell).strip() == "" for cell in row_data[:5]):
                        empty_row_count += 1
                    else:
                        empty_row_count = 0  # Reset if a non-empty row is found

                    # Stop processing if 3 or more consecutive empty rows are found
                    if empty_row_count >= 3:
                        break
                    
                    # Create a dictionary (object) for the row if the first 5 cells are not empty
                    if any(cell is not None and str(cell).strip() != "" for cell in row_data[:5]):  # Only check the first 5 cells
                        row_object = {headers[i]: row_data[i] for i in range(len(headers))}
                        extracted_data.append(row_object)

                # Step 3: Print or return the structured dataset
                data["items"] = format_float_values(extracted_data)  # For demonstration, can be modified as needed

            except Exception as e:
                logging.error(f"Error processing Excel file: {e}")
                return func.HttpResponse(
                    body=json.dumps({"error": f"Failed to process Excel file: {str(e)}"}),
                    status_code=500,
                    mimetype="application/json"
                )

            # Delete the temporary uploaded file
            os.remove(uploaded_file_path)

            # Write the extracted data to an Excel file
            excel_file = write_to_excel(data)
            logging.info("Generated Excel file.")

            reference = data["Reference"] if data["Reference"] else '#No_Ref#'

            # Set response headers for the Excel file download
            headers = {
                'Content-Disposition': 'attachment; filename="' + reference + '.xlsx"',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }

            # Return the Excel file as an HTTP response
            return func.HttpResponse(excel_file.getvalue(), headers=headers, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        except Exception as e:
            logging.error(f"Error decoding or saving file: {e}")
            return func.HttpResponse(
                body=json.dumps({"error": f"Failed to decode base64 file: {str(e)}"}),
                status_code=500,
                mimetype="application/json"
            )

    
    

    

    
    
    
    
    

    
    
