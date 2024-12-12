from io import BytesIO
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Alignment, Border, Side

# Function to generate Excel for each container with styles using openpyxl
def generate_excel_zip(data):
    # Create a BytesIO buffer to hold the ZIP file data
    zip_buffer = BytesIO()

    # Create a ZIP archive in the buffer
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for entry in data:
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Create the header font style
            bold_font = Font(bold=True)
            centered_alignment = Alignment(horizontal='center', vertical='center')
            left_aligned = Alignment(horizontal='left')
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            invoice_number = entry['items'][0]['Invoice_Number']

            # General information with titles in column A and values in column B
            general_info = [
                ('Origin', entry['City']),
                ('Vessel', entry['vissel']),
                ('ETA', ''),
                ('Invoice number', entry['container']),
                ('Invoice Date', entry['Date']),
                ('CT Type', '')
            ]

            # Write general info to cells (starting from row 1)
            for idx, (title, value) in enumerate(general_info, start=1):
                ws[f'A{idx}'] = title
                ws[f'B{idx}'] = value
                ws[f'A{idx}'].font = bold_font
                ws[f'A{idx}'].alignment = left_aligned
                ws[f'B{idx}'].alignment = left_aligned

            # Add a few empty rows after general info
            ws.append([])
            ws.append(['Items'])

            # Define the column headers for the items table
            headers = ['Col #1', 'Col #2', 'Col #3', 'Col #4', 'Col #5', 'Col #6', 'Col #7', 'Col #8', 'Col #9', 'Col #10', 'Col #11', 'Col #12']

            # Write the headers to the sheet starting from row 10
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=9, column=col_num, value=header)
                cell.font = bold_font
                
            item_number = 623456    

            # Write the data rows for each item
            row_num = 10
            for item in entry['items']:
                ws[f'A{row_num}'] = item_number
                ws[f'B{row_num}'] = ""
                ws[f'C{row_num}'] = item.get('hs_code', '')
                ws[f'D{row_num}'] = 0#item.get('pieces', '')
                ws[f'E{row_num}'] = item.get('pckg', '')
                ws[f'F{row_num}'] = item.get('net', '')
                ws[f'G{row_num}'] = item.get('gross', '')
                ws[f'H{row_num}'] = item.get('price', '')
                ws[f'I{row_num}'] = entry['Term'][0]  # Term[0]
                ws[f'J{row_num}'] = entry['Term'][1]  # Term[1]
                ws[f'K{row_num}'] = item.get('Invoice_Number', '')
                ws[f'L{row_num}'] = entry['Date']

                row_num += 1
                item_number += 11

            # Save the file with the container name
            excel_filename = f"{invoice_number}-{entry['container']}.pdf.xlsx"
            with BytesIO() as excel_buffer:
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                zip_file.writestr(excel_filename, excel_buffer.read())
    # Finalize the ZIP file and prepare it for the response
    zip_buffer.seek(0)  # Go to the beginning of the ZIP buffer
    return zip_buffer.getvalue()            
